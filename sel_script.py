###### THIS SCRIPT IS FOR SCRAPING THE OLD DATA FROM THE SITE #######
### TODO It cannot scrape the data from different month from what is rendered on the page you can do it easily by clicking on back arrow using selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re
import time

# URL of the webpage to scrape
url = 'https://baseballmonster.com/boxscores.aspx'

# Set up the Selenium WebDriver (this example uses Chrome)
driver = webdriver.Chrome()

# Create a new Excel workbook and add sheets for batting and pitching
wb = Workbook()
ws_bat = wb.create_sheet(title='Batting')
ws_pitch = wb.create_sheet(title='Pitching')

# Function to write match data to a worksheet
def write_match_data(ws, match_data, headers,date):
    team_data = []  # Initialize team_data within the function
    # ws.append(headers)
    for row in match_data:
            if row['Name'] == 'Totals':
                for player in team_data:
                    ws.append([date]+ [player.get(header, '') for header in headers])
                # ws.append([])  # Add an empty row to separate teams
                team_data = []  # Reset team_data for the next team
            else:
                team_data.append(row)
    if team_data:
            for player in team_data:
                ws.append([player.get(header, '') for header in headers])
            # ws.append([])  # Add an empty row to separate teams
    # ws.append([])  # Add another empty row to separate matches

try:
    # Open the web page
    driver.get(url)
    wait = WebDriverWait(driver, 40)
    wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_Calendar1")))
    date_neter=[]
    date_nete=""
    # Iterate over each anchor element
    for i in range(40):
        table = driver.find_element(By.ID, "ContentPlaceHolder1_Calendar1")
        anchors = table.find_elements(By.TAG_NAME, "a")

        if i >= len(anchors):
            break

        wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_Calendar1")))
        date = anchors[i].text
        date_neter.append(date + 'May 2024')
        date_nete = date + 'March 2024'
        print(f"Clicking on anchor with date: {date}")
        anchors[i].click()


        # Get the page source after JavaScript has rendered the content
        html_content = driver.page_source

        soup = BeautifulSoup(html_content, 'html.parser')

        # Find all tables with the class 'datatable w3-hoverable'
        tables = soup.find_all('table', {'class': 'datatable w3-hoverable'})

        # Find match details
        find_teams = soup.find_all('p', {'class': 'text minP'})
        match_between_teams = [match.text for match in find_teams]

        # Initialize lists to store the extracted data from all tables
        matches_data = []

        # Process tables in pairs
        for j in range(0, len(tables), 2):
            if j + 1 >= len(tables):
                break

            table_bat = tables[j]
            table_pitch = tables[j + 1]

            # Initialize lists to store the extracted data for the current match
            match_data_bat = []
            match_data_pitch = []

            # Process batting table
            thead_bat = table_bat.find('thead')
            if thead_bat:
                headers_bat = [header.text.strip() for header in thead_bat.find_all('th')]
                headers_bat.append('Opponent')
                headers_bat.append('Win/Loss')
                rows_bat = table_bat.find_all('tr')[1:]  # Skipping the first row which is the header
                for row in rows_bat:
                    cols_bat = row.find_all('td')
                    cols_bat = [col.text.strip() for col in cols_bat]
                    if not any(cols_bat):
                        continue
                    # Add opponent and win/loss information
                    match_str = match_between_teams[j//2]
                    regex = r"(\b[A-Z]{2,3}\b) \((\d+)\) @ (\b[A-Z]{2,3}\b) \((\d+)\)"
                    regex1 = r'([A-Z]{3}) \((\d+)\) @ ([A-Z]{3}) (\d+) \((\d+)\)'
                    match = re.search(regex, match_str)
                    match1 = re.search(regex1, match_str)
                    if match or match1:
                        if match:
                            team1, score1, team2, score2 = match.groups()
                            score1, score2 = int(score1), int(score2)
                            if cols_bat[6] == team1:
                                opponent = team2
                                win_loss = 'W' if score1 > score2 else 'L'
                            elif cols_bat[6] == team2:
                                opponent = team1
                                win_loss = 'W' if score2 > score1 else 'L'
                            else:
                                opponent = '-'
                                win_loss = '-'
                            cols_bat.append(opponent)
                            cols_bat.append(win_loss)
                        else:
                            team1, score1, team2, score2, mis = match1.groups()
                            score1, score2 = int(score1), int(score2)
                            if cols_bat[6] == team1:
                                opponent = team2
                                win_loss = 'W' if score1 > score2 else 'L'
                            elif cols_bat[6] == team2:
                                opponent = team1
                                win_loss = 'W' if score2 > score1 else 'L'
                            else:
                                opponent = '-'
                                win_loss = '-'
                            cols_bat.append(opponent)
                            cols_bat.append(win_loss)
                    headers_bat[0] = date_neter[i]
                    player_data_bat = dict(zip(headers_bat, cols_bat))
                    match_data_bat.append(player_data_bat)

            # Process pitching table
            thead_pitch = table_pitch.find('thead')
            if thead_pitch:
                headers_pitch = [header.text.strip() for header in thead_pitch.find_all('th')]
                headers_pitch.append('Opponent')
                headers_pitch.append('Win/Loss')
                rows_pitch = table_pitch.find_all('tr')[1:]  # Skipping the first row which is the header
                for row in rows_pitch:
                    cols_pitch = row.find_all('td')
                    cols_pitch = [col.text.strip() for col in cols_pitch]
                    if not any(cols_pitch):
                        continue
                    # Add opponent and win/loss information
                    match_str = match_between_teams[j//2]
                    regex = r"(\b[A-Z]{2,3}\b) \((\d+)\) @ (\b[A-Z]{2,3}\b) \((\d+)\)"
                    regex1 = r'([A-Z]{3}) \((\d+)\) @ ([A-Z]{3}) (\d+) \((\d+)\)'
                    match = re.search(regex, match_str)
                    match1 = re.search(regex1, match_str)
                    if match or match1:
                        if match:
                            team1, score1, team2, score2 = match.groups()
                            score1, score2 = int(score1), int(score2)
                            if cols_pitch[6] == team1:
                                opponent = team2
                                win_loss = 'W' if score1 > score2 else 'L'
                            elif cols_pitch[6] == team2:
                                opponent = team1
                                win_loss = 'W' if score2 > score1 else 'L'
                            else:
                                opponent = '-'
                                win_loss = '-'
                            cols_pitch.append(opponent)
                            cols_pitch.append(win_loss)
                        else:
                            team1, score1, team2, score2, mis = match1.groups()
                            score1, score2 = int(score1), int(score2)
                            if cols_pitch[6] == team1:
                                opponent = team2
                                win_loss = 'W' if score1 > score2 else 'L'
                            elif cols_pitch[6] == team2:
                                opponent = team1
                                win_loss = 'W' if score2 > score1 else 'L'
                            else:
                                opponent = '-'
                                win_loss = '-'
                            cols_pitch.append(opponent)
                            cols_pitch.append(win_loss)
                    headers_pitch[0] = 'Date'
                    player_data_pitch = dict(zip(headers_pitch, cols_pitch))
                    match_data_pitch.append(player_data_pitch)

            # Append match data to the matches data list
            matches_data.append({
                'batting': {
                    'headers': headers_bat,
                    'data': match_data_bat
                },
                'pitching': {
                    'headers': headers_pitch,
                    'data': match_data_pitch
                }
            })

        # Write the batting data to the 'Batting' sheet
        for match in matches_data:
            write_match_data(ws_bat, match['batting']['data'], match['batting']['headers'], date_nete)

        # Write the pitching data to the 'Pitching' sheet
        for match in matches_data:
            write_match_data(ws_pitch, match['pitching']['data'], match['pitching']['headers'], date_nete)

        # Remove the default sheet created by openpyxl
        # del wb['Sheet']
        
        # Save the workbook to a file
        wb.save('baseball_data_newMarch.xlsx')

        # Go back to the calendar page to click the next date
        # driver.back()
        wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_Calendar1")))

finally:
    # Close the WebDriver
    driver.quit()
